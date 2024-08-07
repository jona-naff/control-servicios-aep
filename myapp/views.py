from django.http import JsonResponse, FileResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.shortcuts import render, redirect
from .models import  Avaluos, Clientes, Estados, Municipios, Colonias, Tipos, Valuadores, Estatus, Comentarios, Honorarios, Comentarios, Honorarios,Tiposimb
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import *
import decimal
import csv
import datetime
from django.db.models import Q
import io
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
from reportlab.lib import pagesizes
from django.core.serializers import serialize
from reportlab.platypus import PageBreak
import json

from .forms import AvaluoForm, ColoniaForm, ComentariosForm, HonorariosForm

from django.core.paginator import Paginator 

import io
from django.http import HttpResponse
from django.views.generic import View
import os

from django.http.response import HttpResponse

from xlsxwriter.workbook import Workbook


from django.views import View
from django.conf import settings
from reportlab.lib.utils import ImageReader
from reportlab.lib.colors import Color
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, KeepTogether, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


from datetime import date

# Create your views here.


def formato_fechas(fecha):
    fecha_form = fecha[8:10] + '/' + fecha[5:7] + '/' + fecha[0:4]
    return fecha_form

def formato_fechas_inverso(fecha):
    fecha_form = fecha[6:10] + '-' + fecha[3:5] + '-' + fecha[0:2]
    return fecha_form

def index(request):
    return render(request,'core/index.html')

@login_required
def servicios(request):
    conteo_tipo_servicio = Tipos.objects.all()
    avaluos_por_mes = Avaluos.objects.raw("select '1' as avaluoid, CASE WHEN MONTH(dtcreate) = 1 THEN 'Enero' WHEN MONTH(dtcreate) = 2 THEN 'Febrero' WHEN MONTH(dtcreate) = 3 THEN 'Marzo' WHEN MONTH(dtcreate) = 4 THEN 'Abril' WHEN MONTH(dtcreate) = 5 THEN 'Mayo' WHEN MONTH(dtcreate) = 6 THEN 'Junio' WHEN MONTH(dtcreate) = 7 THEN 'Julio' WHEN MONTH(dtcreate) = 8 THEN 'Agosto' WHEN MONTH(dtcreate) = 9 THEN 'Septiembre' WHEN MONTH(dtcreate) = 10 THEN 'Octubre' WHEN MONTH(dtcreate) = 11 THEN 'Noviembre' WHEN MONTH(dtcreate) = 12 THEN 'Diciembre' ELSE 'Esto no es un mes' END AS 'mes', year(dtcreate) as anho,count(*) as numero from avaluos where colonia_id<>50144 group by month(dtcreate), year(dtcreate) order by year(dtcreate) desc, month(dtcreate) desc")
    avaluos = Avaluos.objects.order_by('-dtsolicitud').select_related('cliente','tipo','valuador','estatus')
    avaluos_dic = []
    primer_dia_mes = formato_fechas(str(date.today().replace(day=1)))
    for avaluo in avaluos:
        id = str(avaluo.avaluoid)
        cliente = str(avaluo.cliente)
        ubicacion = str(avaluo.calle) 
        dtsolicitud = str(avaluo.dtsolicitud)
        dtvaluador = str(avaluo.dtvaluador)
        dtcliente = str(avaluo.dtcliente)
        dtcobro = str(avaluo.dtcobro)
        dtpago = str(avaluo.dtpago)
        valuador = str(avaluo.valuador)
        estatus = str(avaluo.estatus)
        tipo = str(avaluo.tipo)
        colonia = str(avaluo.colonia.municipio_id)
        avaluos_dic.append({'id': id, 
                            'ubicacion': ubicacion, 
                            'dtsolicitud' : dtsolicitud,
                            'dtvaluador' : dtvaluador,
                            'dtcliente' : dtcliente,
                            'dtcobro' : dtcobro,
                            'dtpago' : dtpago,
                            'cliente': cliente,
                            'valuador':valuador,
                            'estatus': estatus,
                            'tipo': tipo,
                            'colonia':colonia
                            })
    p = Paginator(avaluos_dic,40)
    page = request.GET.get('page')
    avaluos_pag = p.get_page(page)
    nums = []
    
    for i in range(avaluos_pag.paginator.num_pages):
        nums.append(i)
    return render(request,'core/servicios.html',{
         'conteo_tipo_servicio': conteo_tipo_servicio,
         'avaluos_por_mes': avaluos_por_mes,
         'avaluos_pag': avaluos_pag,
         'nums': nums,
         'primer_dia_mes': primer_dia_mes
    })
   #return render(request,'core/servicios.html')


def exit(request):
    logout(request)
    return redirect('index')


def get_clientes(request):
    clientes=list(Clientes.objects.values().order_by('nombre'))
    for i in range(len(clientes)):
        if clientes[i]['clienteid'] == 0:
            del clientes[i]
            break
    clientes = [list(Clientes.objects.values())[0]]+clientes
    
    if (len(clientes)>0):
        data={'message':"Success",'clientes':clientes}
    else:
        data={'message': "Not Found"}
        
    return JsonResponse(data)
  
def get_tiposimb(request):
    #tiposimb=list(Tiposimb.objects.values())
    tiposimb=list(Tiposimb.objects.values().order_by('nombre'))
    for i in range(len(tiposimb)):
        if tiposimb[i]['tipoimbid'] == 0:
            del tiposimb[i]
            break
    tiposimb = [list(Tiposimb.objects.values())[0]]+tiposimb
    if (len(tiposimb)>0):
        data={'message':"Success",'tiposimb': tiposimb}
    else:
        data={'message': "Not Found"}
        
    return JsonResponse(data)


    
def get_tipos(request):
    tipos=list(Tipos.objects.values().order_by('display'))
    for i in range(len(tipos)):
        if tipos[i]['tipoid'] == 0:
            del tipos[i]
            break
    tipos = [list(Tipos.objects.values())[0]]+tipos
    if (len(tipos)>0):
        data={'message':"Success",'tipos': tipos}
    else:
        data={'message': "Not Found"}
        
    return JsonResponse(data)



def get_valuadores(request):
    valuadores=list(Valuadores.objects.values().order_by('display'))
    for i in range(len(valuadores)):
        if valuadores[i]['valuadorid'] == 0:
            del valuadores[i]
            break
    valuadores = [list(Valuadores.objects.values())[0]]+valuadores
    
    if (len(valuadores)>0):
        data={'message':"Success",'valuadores': valuadores}
    else:
        data={'message': "Not Found"}
        
    return JsonResponse(data)


def get_estatus(request):
    estatus=list(Estatus.objects.values().order_by('nombre'))
    for i in range(len(estatus)):
        if estatus[i]['estatusid'] == 0:
            del estatus[i]
            break
    estatus = [list(Estatus.objects.values())[0]]+estatus
    
    if (len(estatus)>0):
        data={'message':"Success",'estatus': estatus}
    else:
        data={'message': "Not Found"}
        
    return JsonResponse(data)
   


def get_estados(request):
    estados=list(Estados.objects.values())
    
    if (len(estados)>0):
        data={'message':"Success",'estados':estados}
    else:
        data={'message': "Not Found"}
        
    return JsonResponse(data)


def get_municipios(request, estado_id):
    if estado_id < 1:
        municipios = list(Municipios.objects.values())
    else:
        municipios = [list(Municipios.objects.values())[0]]
        municipios += list(Municipios.objects.order_by('nombre').filter(estado_id=estado_id).values())
    
    if (len(municipios)>1):
            data={'message':"Success",'municipios': municipios}
    else:
        data={'message': "Not Found"}
        
    return JsonResponse(data)


def get_colonias(request, municipio_id):
    if municipio_id<1:
        colonias = list(Colonias.objects.values())[0:200]
    else:
        colonias = [list(Colonias.objects.values())[0]]
        colonias += list(Colonias.objects.order_by('nombre').filter(municipio_id=municipio_id).values())
    
    if (len(colonias)>1):
            data={'message':"Success",'colonias': colonias}
    else:
        data={'message': "Not Found"}
        
    return JsonResponse(data)


def get_tipo_byid(request,tipoid):
    tipo = Tipos.objects.get(tipoid=tipoid)
    display = tipo.display
    tipoid = tipo.tipoid

    if tipo:
        data={'message':"Success", 'tipo': display, 'tipoid': tipoid}
    else:
        data={'message': "Not Found"}

    return JsonResponse(data)


def get_avaluos_inic(request):
    avaluos = Avaluos.objects.order_by('-dtsolicitud').select_related('cliente','tipo','valuador','estatus')
    avaluos_dic = []
    for avaluo in avaluos:
        id = str(avaluo.avaluoid)
        cliente = str(avaluo.cliente)
        ubicacion = str(avaluo.calle) 
        dtsolicitud = str(avaluo.dtsolicitud)
        dtvaluador = str(avaluo.dtvaluador)
        dtcliente = str(avaluo.dtcliente)
        dtcobro = str(avaluo.dtcobro)
        dtpago = str(avaluo.dtpago)
        valuador = str(avaluo.valuador)
        estatus = str(avaluo.estatus)
        tipo = str(avaluo.tipo)
        colonia = str(avaluo.colonia.municipio_id)
        avaluos_dic.append({'id': id, 
                            'ubicacion': ubicacion, 
                            'dtsolicitud' : dtsolicitud,
                            'dtvaluador' : dtvaluador,
                            'dtcliente' : dtcliente,
                            'dtcobro' : dtcobro,
                            'dtpago' : dtpago,
                            'cliente': cliente,
                            'valuador':valuador,
                            'estatus': estatus,
                            'tipo': tipo,
                            'colonia':colonia
                            })
    if (len(avaluos_dic)>0):
            p = Paginator(avaluos_dic,40)
            page = request.GET.get('page')
            avaluos_pag = p.get_page(page)
            nums = []
            for i in range(avaluos_pag.paginator.num_pages):
                nums.append(i)
            has_previous = avaluos_pag.has_previous()
            paginator_info = {
                'current_page': avaluos_pag.number,
                'num_pages': avaluos_pag.paginator.num_pages,
                'per_page': avaluos_pag.paginator.per_page,
                'count': avaluos_pag.paginator.count,
                'has_previous' : avaluos_pag.has_previous(),
                'has_next' : avaluos_pag.has_next(),
                'nums':nums,
            }

            data={'message':"Success",'paginator_info': paginator_info, "avaluos": avaluos_dic}
            
    else:
        data={'message': "Not Found"}
        
    #return render(request, 'core/paginacion.html', {'avaluos': avaluos_dic,'avaluos_pag': avaluos_pag,'nums':nums,'has_previous': avaluos_pag.has_previous(),'paginator_info': paginator_info})
    return JsonResponse(data)

def reduce(function, iterable):
    it = iter(iterable)
    if len(list(iterable))>0:
        value = next(it,Q())
        for element in it:
            value = function(value, element)
    else:
        value = Q()
    return value

def operator_and(val1,val2):
    return val1 & val2

def operator_or(val1,val2):
    return val1 | val2



def get_avaluo(request, avaluoid):
    avaluo = Avaluos.objects.raw("SELECT *,  m.nombre as municipio, c.nombre as colo, e.nombre as estado, t.descripcion as tip, cl.nombre as client, va.display as valua, t.descripcion as tipo_desc, es.nombre as est FROM avaluos as a INNER JOIN colonias AS c on c.colonia_id = a.colonia_id INNER JOIN municipios AS m on m.municipio_id = c.municipio_id INNER JOIN estados AS e on e.estado_id = m.estado_id INNER JOIN tipos AS t on t.tipoid = a.tipo_id  INNER JOIN clientes AS cl on cl.clienteid = a.cliente_id INNER JOIN valuadores AS va on va.valuadorid = a.valuador_id INNER JOIN estatus AS es on es.estatusid = a.estatus_id where avaluoid = %s; ", [avaluoid])
    
    avaluo = Avaluos.objects.filter(avaluoid=avaluoid)
   
    comentarios = Comentarios.objects.raw("Select * from comentarios where avaluo_id = %s", [avaluoid])

    honorarios = Honorarios.objects.raw("Select * from honorarios where avaluo_id = %s", [avaluoid])

    honorarios_dict = []
    for hon in honorarios:
        if hon.monto:
            hon.monto = '{:,}'.format(hon.monto)
        honorarios_dict.append({'razon': hon.razon, 'monto': hon.monto})

    
    av = Avaluos.objects.get(avaluoid=avaluoid)
    colonia_id = av.colonia_id
    if av.consecutivo:
        folio = av.tipo.display  + '-' + av.cliente.nombre + '/' + av.dtcreate[5:7] + '-' + av.dtcreate[2:4] + '/' + str(av.consecutivo) + '-' + av.valuador.display
    else:
        folio = av.tipo.display  + '-' + av.cliente.nombre + '/' + av.dtcreate[5:7] + '-' + av.dtcreate[2:4] + '/' + 'None' + '-' + av.valuador.display

    indicador_tipo = 0
    avaluoid=0
    for av in avaluo:
        
        if av.m2c:
            av.m2c = '{:,}'.format(av.m2c)
        if av.m2t:
            av.m2t = '{:,}'.format(av.m2t)
        if av.tipo.display == "SUP" or av.tipo.display == "VER":
            indicador_tipo = 1

        avaluoid=av.avaluoid
            

    municipio_id = Colonias.objects.get(colonia_id=colonia_id)
    municipio_id = municipio_id.municipio_id
    municipio = Municipios.objects.get(municipio_id=municipio_id)
    estado_id = municipio.estado_id
    municipio=municipio.nombre
    
    estado = Estados.objects.get(estado_id=estado_id)
    estado=estado.nombre
    if request.method == 'GET': 
        
        if indicador_tipo == 1:
            return render(request,'core/detalles.html',{
            'avaluo': avaluo,
            'avaluoid': avaluoid,
            'estado':estado,
            'municipio':municipio,
            'comentarios': comentarios,
            'honorarios': honorarios_dict,
            'folio': folio,
            'indicador':indicador_tipo,
            'AvaluoForm': AvaluoForm,
            'ComentariosForm': ComentariosForm,
            'HonorariosForm': HonorariosForm
        })
        else:
            return render(request,'core/detalles.html',{
            'avaluo': avaluo,
            'estado':estado,
            'municipio':municipio,
            'comentarios': comentarios,
            'honorarios': honorarios,
            'folio': folio,
            'AvaluoForm': AvaluoForm,
            'ComentariosForm': ComentariosForm,
            'HonorariosForm': HonorariosForm
        })


    else: 
        try:
            print(request.POST)
            avaluo_editar = Avaluos.objects.get(avaluoid=avaluoid)
            request.POST = request.POST.copy()
            avaluo_editar.tipo=Tipos.objects.get(tipoid = request.POST.get("tipo",avaluo_editar.tipo.tipoid))
            
            avaluo_editar.nofactura=request.POST.get("nofactura",avaluo_editar.nofactura)
            avaluo_editar.tipoimb=Tiposimb.objects.get(tipoimbid = request.POST.get("tipoimb",avaluo_editar.tipoimb.tipoimbid))
            print(request.POST.get("valuador",avaluo_editar.valuador))
            valuador = request.POST.get("valuador",avaluo_editar.valuador)
            if "valuador" in request.POST:
                avaluo_editar.valuador=Valuadores.objects.get(valuadorid = request.POST.get("valuador",avaluo_editar.valuador))
            
            #if avaluo_editar.valor:
            #    avaluo_editar.valor = float('{:,}'.format(int(avaluo_editar.valor)))

            avaluo_editar.calle=request.POST.get("calle",avaluo_editar.calle)
            avaluo_editar.numero=request.POST.get("numero",avaluo_editar.numero)
            
            #elif not isinstance(avaluo_editar.numeroint, int):
            #    avaluo_editar.numeroint=0
            avaluo_editar.manzana=request.POST.get("manzana",avaluo_editar.manzana)
            avaluo_editar.lote=request.POST.get("lote",avaluo_editar.lote)
            avaluo_editar.colonia=Colonias.objects.get(colonia_id = request.POST.get("colonia",avaluo_editar.colonia.colonia_id))
            avaluo_editar.dictamen=request.POST.get("dictamen",avaluo_editar.dictamen)
            avaluo_editar.proyecto=request.POST.get("proyecto",avaluo_editar.proyecto)
            if "valuador" in request.POST:
                avaluo_editar.valuador=Valuadores.objects.get(valuadorid = request.POST.get("valuador",avaluo_editar.valuador))
            if "m2c" in request.POST:
                request.POST["m2c"] = float(quitar_comas(request.POST["m2c"]))
                avaluo_editar.m2c=request.POST.get("m2c",avaluo_editar.m2c)

            if "m2t" in request.POST:
                request.POST["m2t"] = float(quitar_comas(request.POST["m2t"]))
                avaluo_editar.m2t=request.POST.get("m2t",avaluo_editar.m2t)

            if "valor" in request.POST:
                request.POST["valor"] = float(quitar_comas(request.POST["valor"]))
                avaluo_editar.valor=request.POST.get("valor",avaluo_editar.valor)
                
            if "numeroint" in request.POST:
                avaluo_editar.numeroint=request.POST.get("numeroint",0)
                
                avaluo_editar.save(update_fields=['tipo','m2c','m2t','nofactura','tipoimb','valor','calle','numero','numeroint','manzana','lote','colonia','valuador','dictamen','proyecto'])
            else:
                avaluo_editar.save(update_fields=['tipo','m2c','m2t','nofactura','tipoimb','valor','calle','numero','manzana','lote','colonia','valuador','dictamen','proyecto'])
          
            
            
            today = date.today()
            var_coms = int(request.POST.get("control_comentarios",0))
            var_hons = int(request.POST.get("control_honorarios",0))
            for i in range(0,var_coms):
                num = str(i+1)
                
                coms = Comentarios(comentario=request.POST["comentario_"+num],avaluo_id=avaluoid,fecha=today)
                coms.save()
                
            for i in range(0,var_hons):
                num = str(i+1)
                request.POST["monto_"+num] = float(quitar_comas(request.POST["monto_"+num]))
                hons = Honorarios(razon=request.POST["razon_"+num],monto=request.POST["monto_"+num],avaluo_id=avaluoid)
                hons.save()
                

            comentarios = Comentarios.objects.raw("Select * from comentarios where avaluo_id = %s", [avaluoid])
            honorarios = Honorarios.objects.raw("Select * from honorarios where avaluo_id = %s", [avaluoid])
            for hon in honorarios:
                    if hon.monto:
                        print(hon.monto)
                        hon.monto = '{:,}'.format(hon.monto)
                    else:
                        hon.monto = 0

            avaluo = Avaluos.objects.raw("SELECT *,  m.nombre as municipio, c.nombre as colo, e.nombre as estado, t.descripcion as tip, cl.nombre as client, va.display as valua, t.descripcion as tipo_desc, es.nombre as est FROM avaluos as a INNER JOIN colonias AS c on c.colonia_id = a.colonia_id INNER JOIN municipios AS m on m.municipio_id = c.municipio_id INNER JOIN estados AS e on e.estado_id = m.estado_id INNER JOIN tipos AS t on t.tipoid = a.tipo_id  INNER JOIN clientes AS cl on cl.clienteid = a.cliente_id INNER JOIN valuadores AS va on va.valuadorid = a.valuador_id INNER JOIN estatus AS es on es.estatusid = a.estatus_id where avaluoid = %s; ", [avaluoid])
            avaluo = Avaluos.objects.filter(avaluoid=avaluoid)
            

            #if "valor" in request.POST:
            for av in avaluo:
                if av.valor:
                    print(av.valor)
                    av.valor = '{:,}'.format(av.valor)
                else:
                        av.valor = 0
            av = Avaluos.objects.get(avaluoid=avaluoid)
            if av.consecutivo:
                folio = av.tipo.display  + '-' + av.cliente.nombre + '/' + av.dtcreate[5:7] + '-' + av.dtcreate[2:4] + '/' + str(av.consecutivo) + '-' + av.valuador.display
            else:
                folio = av.tipo.display  + '-' + av.cliente.nombre + '/' + av.dtcreate[5:7] + '-' + av.dtcreate[2:4] + '/' + 'None' + '-' + av.valuador.display
            

            indicador_tipo = 0
            for av in avaluo:

                if av.m2c:
                    av.m2c = '{:,}'.format(av.m2c)
                if av.m2t:
                    av.m2t = '{:,}'.format(av.m2t)
                if av.tipo.display == "SUP" or av.tipo.display == "VER":
                    indicador_tipo = 1
                    
            municipio_id = Colonias.objects.get(colonia_id=colonia_id)
            municipio_id = municipio_id.municipio_id
            municipio = Municipios.objects.get(municipio_id=municipio_id)
            estado_id = municipio.estado_id
            municipio=municipio.nombre
            
            estado = Estados.objects.get(estado_id=estado_id)
            estado=estado.nombre
            
            if indicador_tipo == 1:
                return render(request,'core/detalles.html',{
                'avaluo': avaluo,
                'estado':estado,
                'municipio':municipio,
                'comentarios': comentarios,
                'honorarios': honorarios,
                'folio': folio,
                'indicador':indicador_tipo,
                'AvaluoForm': AvaluoForm,
                'ComentariosForm': ComentariosForm,
                'HonorariosForm': HonorariosForm
            })
            else:
                return render(request,'core/detalles.html',{
                'avaluo': avaluo,
                'estado':estado,
                'municipio':municipio,
                'comentarios': comentarios,
                'honorarios': honorarios,
                'folio': folio,
                'AvaluoForm': AvaluoForm,
                'ComentariosForm': ComentariosForm,
                'HonorariosForm': HonorariosForm
            })
        except Avaluos.DoesNotExist:
            return render(request, 'core/detalles.html',{
            'AvaluoForm': AvaluoForm,
            'ComentariosForm': ComentariosForm,
            'HonorariosForm': HonorariosForm,
            'error': 'Provea información válida'
                })


def get_details(request):
    avaluoid=Avaluos.objects.last().avaluoid
    avaluo = Avaluos.objects.raw("SELECT *,  m.nombre as municipio, c.nombre as colo, e.nombre as estado, t.descripcion as tip, cl.nombre as client, va.display as valua, t.descripcion as tipo_desc, es.nombre as est FROM avaluos as a INNER JOIN colonias AS c on c.colonia_id = a.colonia_id INNER JOIN municipios AS m on m.municipio_id = c.municipio_id INNER JOIN estados AS e on e.estado_id = m.estado_id INNER JOIN tipos AS t on t.tipoid = a.tipo_id  INNER JOIN clientes AS cl on cl.clienteid = a.cliente_id INNER JOIN valuadores AS va on va.valuadorid = a.valuador_id INNER JOIN estatus AS es on es.estatusid = a.estatus_id where avaluoid = %s; ", [avaluoid])
    av = Avaluos.objects.get(avaluoid=avaluoid)
    if av.consecutivo:
        folio = av.tipo.display  + '-' + av.cliente.nombre + '/' + av.dtcreate[5:7] + '-' + av.dtcreate[2:4] + '/' + str(av.consecutivo) + '-' + av.valuador.display
    else:
        folio = av.tipo.display  + '-' + av.cliente.nombre + '/' + av.dtcreate[5:7] + '-' + av.dtcreate[2:4] + '/' + 'None' + '-' + av.valuador.display

    avaluo = Avaluos.objects.filter(avaluoid=avaluoid)
    colonia_id = av.colonia_id
    comentarios = Comentarios.objects.raw("Select * from comentarios where avaluo_id = %s", [avaluoid])

    honorarios = Honorarios.objects.raw("Select * from honorarios where avaluo_id = %s", [avaluoid])
    #av = Avaluos.objects.get(avaluoid=avaluoid)
    indicador_tipo = 0
    for av in avaluo:
        if av.valor:
            av.valor = '{:,}'.format(av.valor)
            
        else:
            av.valor = 0
    for av in avaluo:
        if av.m2c:
            av.m2c = '{:,}'.format(av.m2c)
        if av.m2t:
            av.m2t = '{:,}'.format(av.m2t)
        if av.tipo.display == "SUP" or av.tipo.display == "VER":
            indicador_tipo = 1
        
        

    municipio_id = Colonias.objects.get(colonia_id=colonia_id)
    municipio_id = municipio_id.municipio_id
    municipio = Municipios.objects.get(municipio_id=municipio_id)
    estado_id = municipio.estado_id
    municipio=municipio.nombre
    
    estado = Estados.objects.get(estado_id=estado_id)
    estado=estado.nombre
         
    
    if indicador_tipo == 1:
        return render(request,'core/detalles_mid.html',{
        'avaluo': avaluo,
        'estado': estado,
        'municipio': municipio,
        'comentarios': comentarios,
        'honorarios': honorarios,
        'folio': folio,
        'indicador':indicador_tipo,
        'AvaluoForm': AvaluoForm,
        'ComentariosForm': ComentariosForm,
        'HonorariosForm': HonorariosForm
    })
    else:
        return render(request,'core/detalles_mid.html',{
        'avaluo': avaluo,
        'comentarios': comentarios,
        'estado': estado,
        'municipio': municipio,
        'honorarios': honorarios,
        'folio': folio,
        'AvaluoForm': AvaluoForm,
        'ComentariosForm': ComentariosForm,
        'HonorariosForm': HonorariosForm
    })



def query_by_id(cliente_id, tipo_id, valuador_id, estatus_id, estado_id, municipio_id, colonia_id):

    ids=[]

    data = [cliente_id, tipo_id, valuador_id, estatus_id, estado_id, municipio_id, colonia_id]
    compare = [0,0,0,0,0,0]
    
    if data == compare:
        var = Q()
    else:
        if cliente_id != 0:
            ids.append(Q(cliente_id=cliente_id))
        if tipo_id != 0:
            ids.append(Q(tipo_id=tipo_id))
        if valuador_id != 0:
            ids.append(Q(valuador_id=valuador_id))
        if estatus_id != 0:
            ids.append(Q(estatus_id=estatus_id))
        if colonia_id != 0:
            ids.append(Q(colonia_id=colonia_id))
        if municipio_id != 0:
            ids.append(Q(colonia__municipio_id=municipio_id))

        ids_or = []
        if estado_id != 0:
            municipios = list(Municipios.objects.filter(estado_id=estado_id).values())
            if len(municipios)>0:
                for municipio in municipios:
                    municipio_id_loc = municipio["municipio_id"]
                    ids_or.append(Q(colonia__municipio_id=municipio_id_loc))
                    
            else:
                pass
        var = reduce(operator_and,ids) & reduce(operator_or,ids_or)
    return var
    

def query_by_date(dtcreate_inicial, dtcreate_final, dtsolicitud_inicial, dtsolicitud_final, dtvaluador_inicial, dtvaluador_final, dtcliente_inicial, dtcliente_final, dtcobro_inicial, dtcobro_final, dtpago_inicial, dtpago_final):
    
    create_range = Q()
    solicitud_range =  Q()
    valuador_range =  Q()
    cliente_range =  Q()
    cobro_range =  Q()
    pago_range = Q()


    if (dtcreate_inicial != '0000-00-00' and dtcreate_inicial != ''):
        create_range &= Q(dtcreate__range=[dtcreate_inicial,'2100-10-10'])

    if (dtcreate_final != '0000-00-00' and dtcreate_final != ''):
        create_range &= Q(dtcreate__range=['1890-10-10',dtcreate_final])


    if (dtsolicitud_inicial != '0000-00-00' and dtsolicitud_inicial != ''):
        solicitud_range &= Q(dtsolicitud__range=[dtsolicitud_inicial,'2100-10-10'])

    if (dtsolicitud_final != '0000-00-00' and dtsolicitud_final != ''):
        solicitud_range &= Q(dtsolicitud__range=['1890-10-10',dtsolicitud_final])

    
    if (dtvaluador_inicial != '0000-00-00' and dtvaluador_inicial != ''):
        valuador_range &= Q(dtvaluador__range=[dtvaluador_inicial,'2100-10-10'])

    if (dtvaluador_final != '0000-00-00' and dtvaluador_final != ''):
        valuador_range &= Q(dtvaluador__range=['1890-10-10',dtvaluador_final])


    if (dtcliente_inicial != '0000-00-00' and dtcliente_inicial != ''):
        cliente_range &= Q(dtcliente__range=[dtcliente_inicial,'2100-10-10'])

    if (dtcliente_final != '0000-00-00' and dtcliente_final != ''):
        cliente_range &= Q(dtcliente__range=['1890-10-10',dtcliente_final])


    if (dtcobro_inicial != '0000-00-00' and dtcobro_inicial != ''):
        cobro_range &= Q(dtcobro__range=[dtcobro_inicial,'2100-10-10'])

    if (dtcobro_final != '0000-00-00' and dtcobro_final != ''):
        cobro_range &= Q(dtcobro__range=['1890-10-10',dtcobro_final])


    if (dtpago_inicial != '0000-00-00' and dtpago_inicial != ''):
        pago_range &= Q(dtpago__range=[dtpago_inicial,'2100-10-10'])

    if (dtpago_final != '0000-00-00' and dtpago_final != ''):
        pago_range &= Q(dtpago__range=['1890-10-10',dtpago_final])


    return create_range & solicitud_range & valuador_range & cliente_range & cobro_range & pago_range
    


def get_avaluos(request, cliente_id, tipo_id, valuador_id, estatus_id, estado_id, municipio_id, colonia_id,dtcreate_inicial, dtcreate_final,dtsolicitud_inicial, dtsolicitud_final, dtvaluador_inicial, dtvaluador_final, dtcliente_inicial, dtcliente_final, dtcobro_inicial, dtcobro_final,dtpago_inicial, dtpago_final):

    query_id = query_by_id(cliente_id, tipo_id, valuador_id, estatus_id, estado_id, municipio_id, colonia_id)
    query_date = query_by_date(dtcreate_inicial, dtcreate_final,dtsolicitud_inicial, dtsolicitud_final, dtvaluador_inicial, dtvaluador_final, dtcliente_inicial, dtcliente_final, dtcobro_inicial, dtcobro_final,dtpago_inicial, dtpago_final)
    avaluos = Avaluos.objects.filter(query_id & query_date).order_by('-avaluoid').select_related('cliente','tipo','valuador','estatus')

    avaluos_tot = Avaluos.objects.values()
    count_avaluos = 1
    for avaluo in avaluos_tot:
        count_avaluos += 1
    
    
    avaluos_dic = []
    
    for avaluo in avaluos:
        id = str(avaluo.avaluoid)
        cliente = str(avaluo.cliente)
        municipioid = avaluo.colonia.municipio_id
        municipio = Municipios.objects.filter(municipio_id=municipioid)
        municipio_nombre = municipio[0].nombre
        #estadoid = Municipios.objects.filter(municipio_id=municipioid)
        estadoid = municipio[0].estado.estado_id
        estado = Estados.objects.filter(estado_id=estadoid)
        estado_nombre=estado[0].nombre
        colonia = str(avaluo.colonia.nombre)
        calle = str(avaluo.calle) 
        manzana = str(avaluo.manzana)
        lote = str(avaluo.lote)
        numero = str(avaluo.numero)
        dtcreate = str(avaluo.dtcreate)
        dtsolicitud = str(avaluo.dtsolicitud)
        dtvaluador = str(avaluo.dtvaluador)
        dtcliente = str(avaluo.dtcliente)
        dtcobro = str(avaluo.dtcobro)
        dtpago = str(avaluo.dtpago)
        valuador = str(avaluo.valuador)
        estatus = str(avaluo.estatus)
        tipo = str(avaluo.tipo)
        consecutivo = str(avaluo.consecutivo)
        tipoimbid = str(avaluo.tipoimb.nombre)
        valor = str(avaluo.valor)
        dictamen = str(avaluo.dictamen)
        proyecto = str(avaluo.proyecto)
        avaluos_dic.append({'id': id, 
                            'estado': estado_nombre,
                            'municipio': municipio_nombre,
                            'colonia': colonia,
                            'calle': calle, 
                            'manzana': manzana, 
                            'lote': lote, 
                            'numero': numero, 
                            'dtcreate': dtcreate,
                            'dtsolicitud' : dtsolicitud,
                            'dtvaluador' : dtvaluador,
                            'dtcliente' : dtcliente,
                            'dtcobro' : dtcobro,
                            'dtpago' : dtpago,
                            'cliente': cliente,
                            'valuador':valuador,
                            'estatus': estatus,
                            'tipo': tipo,
                            'consecutivo':consecutivo,
                            'tipoimbid':tipoimbid,
                            'valor': valor,
                            'dictamen': dictamen,
                            'proyecto': proyecto
                            })
    
    if (len(avaluos_dic)>0):

            num_pags = len(avaluos_dic)//40 + 1
            num_pags_tot = count_avaluos//40 + 1
            data={'message':"Success",'avaluos': avaluos_dic, 'num_pags': num_pags, 'num_pags_tot': num_pags_tot}
    else:
        data={'message': "Not Found"}
        
    return JsonResponse(data)


def get_avaluos_bymunicipio(request,municipio_id):
    avaluos = Avaluos.objects.filter(colonia__municipio_id=municipio_id).order_by('-dtsolicitud').select_related('cliente','tipo','valuador','estatus')
    avaluos_dic = []
    if len(avaluos)>0:
        for avaluo in avaluos:
            colonia = str(avaluo.colonia.municipio_id)
            avaluos_dic.append({'colonia': colonia})

        if (len(avaluos_dic)>0):
            data={'message':"Success",'avaluos': avaluos_dic}
    else:
        data={'message': "Not Found"}
    return JsonResponse(data)


def get_avaluos_bydate(request, dtsolicitud_inicial, dtsolicitud_final, dtvaluador_inicial, dtvaluador_final, dtcliente_inicial, dtcliente_final, dtcobro_inicial, dtcobro_final):

    solicitud_range =  Q()
    valuador_range =  Q()
    cliente_range =  Q()
    cobro_range =  Q()

    if (dtsolicitud_inicial != '0000-00-00' and dtsolicitud_inicial != '') and (dtsolicitud_final != '0000-00-00' and dtsolicitud_final != ''):
        solicitud_range = Q(dtsolicitud__range=[dtsolicitud_inicial,dtsolicitud_final])

    if (dtvaluador_inicial != '0000-00-00' and dtvaluador_inicial != '') and (dtvaluador_final != '0000-00-00' and dtvaluador_final != ''):
        valuador_range = Q(dtvaluador__range=[dtvaluador_inicial,dtvaluador_final])

    if (dtcliente_inicial != '0000-00-00' and dtcliente_inicial != '') and (dtcliente_final != '0000-00-00' and dtcliente_final != ''):
        cliente_range = Q(dtcliente__range=[dtcliente_inicial,dtcliente_final])

    if (dtcobro_inicial != '0000-00-00' and dtcobro_inicial != '') and (dtcobro_final != '0000-00-00' and dtcobro_final != ''):
        cobro_range = Q(dtcobro__range=[dtcobro_inicial,dtcobro_final])


    avaluos = Avaluos.objects.filter(solicitud_range & valuador_range & cliente_range & cobro_range).order_by('-dtsolicitud').select_related('cliente','tipo','valuador','estatus')
    
    avaluos_dic = []
    for avaluo in avaluos:
        id = str(avaluo.avaluoid)
        cliente = str(avaluo.cliente)
        ubicacion = str(avaluo.calle) 

        dtsolicitud = str(avaluo.dtsolicitud)
        dtvaluador = str(avaluo.dtvaluador)
        dtcliente = str(avaluo.dtcliente)
        dtcobro = str(avaluo.dtcobro)
        dtpago = str(avaluo.dtpago)

        valuador = str(avaluo.valuador)
        estatus = str(avaluo.estatus)
        tipo = str(avaluo.tipo)
        
        avaluos_dic.append({'id': id, 
                            'ubicacion': ubicacion, 
                            'dtsolicitud' : dtsolicitud,
                            'dtvaluador' : dtvaluador,
                            'dtcliente' : dtcliente,
                            'dtcobro' : dtcobro,
                            'dtpago' : dtpago,
                            'cliente': cliente,
                            'valuador':valuador,
                            'estatus': estatus,
                            'tipo': tipo
                            })
        
    if (len(avaluos_dic)>0):
            data={'message':"Success",'avaluos': avaluos_dic}
    else:
        data={'message': "Not Found"}
        
    return JsonResponse(data)


def generar_pdf(request, cliente_id, tipo_id, valuador_id, estatus_id, estado_id, municipio_id, colonia_id,dtsolicitud_inicial, dtsolicitud_final, dtvaluador_inicial, dtvaluador_final, dtcliente_inicial, dtcliente_final, dtcobro_inicial, dtcobro_final):
    #Crear Bytestream buffer
    buf = io.BytesIO()
    # Crear canvas
    c = canvas.Canvas(buf,pagesize=letter, bottomup=0)
    #Create a text object
    textob = c.beginText()
    textob.setTextOrigin(inch,inch)
    textob.setFont("Helvetica", 14)

    data = get_avaluos(request, cliente_id, tipo_id, valuador_id, estatus_id, estado_id, municipio_id, colonia_id,dtsolicitud_inicial, dtsolicitud_final, dtvaluador_inicial, dtvaluador_final, dtcliente_inicial, dtcliente_final, dtcobro_inicial, dtcobro_final)
    data = json.loads(data.content)
    avaluos = data['avaluos']
    lines =[]

    for avaluo in avaluos:
        renglon = str(avaluo['id']) + ' ' 
        renglon += avaluo['ubicacion'] + ' '
        renglon += avaluo['cliente'] + ' '
        renglon += avaluo['valuador'] + ' '
        lines.append(renglon)
    #Loop
    for line in lines:
        textob.textLine(line)

    #Finish Up
    c.drawText(textob)
    c.showPage()
    c.save()
    buf.seek(0)

    #Return stuff
    return FileResponse(buf, as_attachment=True, filename='servicios.pdf')

def max_words_per_line(cell_width, font_size, text):
    """
    Calculate the maximum number of words that can fit in one line of a table cell.

    Parameters:
    - cell_width: Width of the table cell.
    - font_size: Font size of the text.
    - text: The text for which to calculate the maximum words.

    Returns:
    - The maximum number of words that can fit in one line.
    """
    words = text.split()
    word_widths = [len(word) * font_size * 0.4 for word in words]  # 0.6 is an estimation factor

    current_width = 0
    max_words = 0
    max_words_list = []
    amount_words = 0

    for i, word_width in enumerate(word_widths):
        if i == 0:
            max_words_list += [words[i]]
            amount_words += 1   
            #current_width += word_width
        else:
            current_width += word_width
            if current_width <= cell_width:
                max_words += 1
                amount_words += 1 
                if i==len(words)-1:#words[0: amount_words] == words:
                    max_words_list += ['^^'] + words[amount_words-max_words: amount_words+1]

            else:
                #amount_words += 1   
                if max_words > 0:
                    max_words_list += ['^^'] + words[amount_words-max_words: amount_words+1]
                #else:
                #    max_words_list += ['^^'] + [words[i]]
                current_width = 0
                max_words = 0
                


    return max_words_list


class GeneratePDFView(View):

    def get(self, request, cliente_id, tipo_id, valuador_id, estatus_id, estado_id, municipio_id, colonia_id,dtcreate_inicial, dtcreate_final,dtsolicitud_inicial, dtsolicitud_final, dtvaluador_inicial, dtvaluador_final, dtcliente_inicial, dtcliente_final, dtcobro_inicial, dtcobro_final,dtpago_inicial, dtpago_final,*args, **kwargs):
        # Create a response object with PDF content type
        response = HttpResponse(content_type='application/pdf')
        today = str(date.today())
        today = today[2:4]+today[5:7]+today[8:10]
        filename = today+"_reporte_control_servicios.pdf"
        # Set the Content-Disposition header to force download
        response['Content-Disposition'] = 'inline; filename="{}"'.format(filename)

        # Create the PDF using the ReportLab canvas
        p = canvas.Canvas(response, pagesize=letter+(2000,0))

        # Add an image at the beginning of the page
        image_path = os.path.join(settings.STATIC_ROOT, 'imagenes/logo.jpg')
        p.drawImage(image_path, inch - 20, letter[1] -inch -20, width=100, height=70)

        #p.setFont("Helvetica-Bold", 12)

        # Add the text "Hello World" below the image
        #p.drawString(inch + 80, letter[1] - inch + 35, "Avalúos, Proyectos y Servicios")

        p.setFont("Helvetica", 12)

        p.drawString(inch + 80, letter[1] - inch + 15 , "www.aep.com.mx")

        line_start = inch - 20
        line_end = line_start + 510 # Adjust based on text width
        line_y = letter[1] - inch - 20  # Adjust based on the position of the text
        p.line(line_start, line_y, line_end, line_y)

        fill_color_rgb = (207/238, 197/238, 238/238)
        fill_color = Color(*fill_color_rgb)
        p.setFillColor(fill_color)

        # Define the coordinates for the cell
        cell_x = inch -15
        cell_y = letter[1] - inch - 55
        cell_width = 500  # Adjust based on text width
        cell_height = 25  # Adjust based on the desired height

        # Draw a colored cell
        p.rect(cell_x, cell_y, cell_width, cell_height, fill=1,  stroke=0)

        # Set the fill color back to black for the text
        p.setFillColor("black")

        # Add text within the colored cell
        p.drawString(cell_x + 225, cell_y + 8, "Servicios")

        # Move to a new line before adding the table
        p.translate(inch, -2 * inch)
        custom_style = ParagraphStyle(
            'custom_style',
            parent=getSampleStyleSheet()['BodyText'],
            alignment=1,
            fontSize=8,  # Adjust the font size as needed
        )
        estado = Estados.objects.get(estado_id = estado_id)
        estado = Paragraph(estado.nombre, custom_style)

        municipio = Municipios.objects.get(municipio_id = municipio_id)
        municipio = Paragraph(municipio.nombre, custom_style)

        colonia = Colonias.objects.get(colonia_id = colonia_id)
        colonia = Paragraph(colonia.nombre, custom_style)

        tipo = Tipos.objects.get(tipoid = tipo_id)
        tipo = Paragraph(tipo.display, custom_style)

        cliente = Clientes.objects.get(clienteid = cliente_id)
        cliente = Paragraph(cliente.nombre, custom_style)

        valuador = Valuadores.objects.get(valuadorid = valuador_id)
        valuador = Paragraph(valuador.display, custom_style)

        estatus = Estatus.objects.get(estatusid = estatus_id)
        estatus = Paragraph(estatus.nombre, custom_style)
        # Define data for the table
        data = [['Parámetros',''],
                ['Estado', estado],
                ['Municipio', municipio],
                ['Colonia', colonia],
                ['Tipo de servicio', tipo],
                ['Cliente', cliente],
                ['Valuador', valuador],
                ['Estatus', estatus]]

        col_widths = [110, 110]

        # Create the table and set style
        table1 = Table(data,colWidths=col_widths)
        style = [('BACKGROUND', (0, 0), (-1, 0), fill_color),
                            ('TEXTCOLOR', (0, 0), (-1, 0), 'black'),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                            ('INNERGRID', (0, 0), (-1, -1), 0.25, 'black'),
                            ('BOX', (0, 0), (-1, -1), 0.25, 'black'),
                            ('GRID',(0,0),(-1,-1),0.5,colors.black),
                            ('SPAN', (0, 0), (-1, 0))]
        # Add styles for alternating grey colors and border texture
        for i in range(1, len(data)):
            if i % 2 == 1:  # Alternating rows
                style += ([('BACKGROUND', (0, i), (-1, i), Color(0.9, 0.9, 0.9)),
                               ('BOX', (0, i), (-1, i), 0.25, 'black'),
                               ('BOTTOMPADDING', (0, i), (-1, i), 5)])
            if i % 2 == 0:  # Alternating rows
                style += ([('BACKGROUND', (0, i), (-1, i), Color(0.8, 0.8, 0.8)),
                            ('BOX', (0, i), (-1, i), 0.25, 'black'),
                            ('BOTTOMPADDING', (0, i), (-1, i), 5)])
        

        style = TableStyle(style)

        table1.setStyle(style)
        w, h = table1.wrap(0, 0)
        # Draw the table on the canvas
        table1.wrapOn(p, 400,400)
        table1.drawOn(p, inch-75, 797 - h)
        

        #table3.wrapOn(p, 800,800)
        #table3.drawOn(p, inch-130, 610 - h)
        dtcrt_inicial = dtcreate_inicial
        dtcrt_final = dtcreate_final
        dtsol_inicial = dtsolicitud_inicial
        dtsol_final = dtsolicitud_final
        dtval_inicial = dtvaluador_inicial
        dtval_final = dtvaluador_final
        dtclt_inicial = dtcliente_inicial
        dtclt_final = dtcliente_final
        dtcbr_inicial = dtcobro_inicial
        dtcbr_final = dtcobro_final
        dtpg_inicial = dtpago_inicial
        dtpg_final = dtpago_final

        data2 = [['Parámetros','',''],
        ['Fecha','Inicio','Fin'],
        ['Alta', dtcrt_inicial,dtcrt_final],
        ['Solicitud', dtsol_inicial,dtsol_final],
        ['Entrega cliente', dtclt_inicial,dtclt_final],
        ['Entrega valuador', dtval_inicial,dtval_final],
        ['Cobro', dtcbr_inicial,dtcbr_final],
        ['Pago', dtpg_inicial,dtpg_final]]

        data2_dict = {'header1':['Parámetros','',''],
        'header2':['Fecha','Inicio','Fin']}


        log_list = [((dtcreate_inicial == '0000-00-00') or (dtcreate_inicial == '')),((dtcreate_final == '0000-00-00') or (dtcreate_final == '')),((dtsolicitud_inicial == '0000-00-00') or (dtsolicitud_inicial == '')),( (dtsolicitud_final == '0000-00-00') or (dtsolicitud_final == '')),((dtvaluador_inicial == '0000-00-00' or dtvaluador_inicial == '')),((dtvaluador_final == '0000-00-00' or dtvaluador_final == '')),((dtcliente_inicial == '0000-00-00' or dtcliente_inicial == '')),((dtcliente_final == '0000-00-00' or dtcliente_final == '')),((dtcobro_inicial == '0000-00-00' or dtcobro_inicial == ''))and((dtcobro_final == '0000-00-00' or dtcobro_final == '')),((dtpago_inicial == '0000-00-00' or dtpago_inicial == '')),((dtpago_final == '0000-00-00' or dtpago_final == ''))]

        log_val = False

        for x in log_list:
            log_val = log_val or not x
        
        if log_val == True:
            print(not ((dtcreate_inicial == '0000-00-00') or (dtcreate_inicial == '')and(dtcreate_final == '0000-00-00') or (dtcreate_final == '')))
            if not ((dtcreate_inicial == '0000-00-00') or (dtcreate_inicial == '')and(dtcreate_final == '0000-00-00') or (dtcreate_final == '')):
                
                data2_dict['alta'] = ['Alta', '','']
                
                if not ((dtcreate_inicial == '0000-00-00') or (dtcreate_inicial == '')):
                    data2_dict['alta'][1] = formato_fechas(dtcrt_inicial)
                if not ((dtcreate_final == '0000-00-00') or (dtcreate_final == '')):
                    data2_dict['alta'][2] = formato_fechas(dtcrt_final)

            if not ((dtsolicitud_inicial == '0000-00-00') or (dtsolicitud_inicial == '')and(dtsolicitud_final == '0000-00-00') or (dtsolicitud_final == '')):
                data2_dict['solicitud'] = ['Solicitud', '','']
                if not ((dtsolicitud_inicial == '0000-00-00') or (dtsolicitud_inicial == '')):
                    data2_dict['solicitud'][1] = formato_fechas(dtsol_inicial)
                if not ((dtsolicitud_final == '0000-00-00') or (dtsolicitud_final == '')):
                    data2_dict['solicitud'][2] = formato_fechas(dtsol_final)

            if not ((dtvaluador_inicial == '0000-00-00') or (dtvaluador_inicial == '')and(dtvaluador_final == '0000-00-00') or (dtvaluador_final == '')):
                data2_dict['valuador'] = ['Valuador', '','']
                if not ((dtvaluador_inicial == '0000-00-00' or dtvaluador_inicial == '')):
                    data2_dict['valuador'][1] = formato_fechas(dtval_inicial)
                if not ((dtvaluador_final == '0000-00-00' or dtvaluador_final == '')):
                    data2_dict['valuador'][2] = formato_fechas(dtval_final)

            if not ((dtcliente_inicial == '0000-00-00') or (dtcliente_inicial == '')and(dtcliente_final == '0000-00-00') or (dtcliente_final == '')):
                data2_dict['cliente'] = ['Cliente', '','']
                if not ((dtcliente_inicial == '0000-00-00' or dtcliente_inicial == '')):
                    data2_dict['cliente'][1] = formato_fechas(dtclt_inicial)
                if not ((dtcliente_final == '0000-00-00' or dtcliente_final == '')):
                    data2_dict['cliente'][2] = formato_fechas(dtclt_final)

            if not ((dtcobro_inicial == '0000-00-00') or (dtcobro_inicial == '')and(dtcobro_final == '0000-00-00') or (dtcobro_final == '')):
                data2_dict['cobro'] = ['Cobro', '','']
                if not ((dtcobro_inicial == '0000-00-00' or dtcobro_inicial == '')):
                    data2_dict['cobro'][1] = formato_fechas(dtcbr_inicial)
                if not ((dtcobro_final == '0000-00-00' or dtcobro_final == '')):
                    data2_dict['cobro'][2] = formato_fechas(dtcbr_final)

            if not ((dtpago_inicial == '0000-00-00') or (dtpago_inicial == '')and(dtpago_final == '0000-00-00') or (dtpago_final == '')):
                data2_dict['pago'] = ['Pago', '','']
                if not ((dtpago_inicial == '0000-00-00' or dtpago_inicial == '')):
                    data2_dict['pago'][1] = formato_fechas(dtpg_inicial)
                if not ((dtpago_final == '0000-00-00' or dtpago_final == '')):
                    data2_dict['pago'][2] = formato_fechas(dtpg_final)

        
            data2 = []

            for x in data2_dict:
                
                data2.append(data2_dict[x])

         
            col_widths2 = [80, 80]
            table2 = Table(data2,colWidths=col_widths2)
            style = [('BACKGROUND', (0, 0), (-1, 1), fill_color),
                                ('TEXTCOLOR', (0, 0), (-1, 0), 'black'),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                ('INNERGRID', (0, 0), (-1, -1), 0.25, 'black'),
                                ('BOX', (0, 0), (-1, -1), 0.25, 'black'),
                                ('GRID',(0,0),(-1,-1),0.5,colors.black),
                                ('SPAN', (0, 0), (-1, 0))]
            for i in range(2, len(data2)):
                if i % 2 == 1:  # Alternating rows
                    style += ([('BACKGROUND', (0, i), (-1, i), Color(0.9, 0.9, 0.9)),
                                ('BOX', (0, i), (-1, i), 0.25, 'black'),
                                ('BOTTOMPADDING', (0, i), (-1, i), 5)])
                if i % 2 == 0:  # Alternating rows
                    style += ([('BACKGROUND', (0, i), (-1, i), Color(0.8, 0.8, 0.8)),
                                ('BOX', (0, i), (-1, i), 0.25, 'black'),
                                ('BOTTOMPADDING', (0, i), (-1, i), 5)])
                
            style = TableStyle(style)
            table2.setStyle(style)
            w, h = table2.wrap(0, 0)
            # Draw the table on the canvas
            table2.wrapOn(p, 400,400)
            table2.drawOn(p, inch+165, 797-h)
        


        #data3 = self.get_avaluos(request, cliente_id, tipo_id, valuador_id, estatus_id, estado_id, municipio_id, colonia_id)
        avaluos = get_avaluos(request, cliente_id, tipo_id, valuador_id, estatus_id, estado_id, municipio_id, colonia_id,dtcreate_inicial, dtcreate_final,dtsolicitud_inicial, dtsolicitud_final, dtvaluador_inicial, dtvaluador_final, dtcliente_inicial, dtcliente_final, dtcobro_inicial, dtcobro_final,dtpago_inicial, dtpago_final)
        avaluos = json.loads(avaluos.content)
        avaluos = avaluos['avaluos']
        
        data3 = []#[['Id','Dirección', 'Fechas','Cliente','Estatus', 'Folio','Tipo Inmueble','Valor']]

        col_widths3 = [35,105,95,65,62,65,60,70]

        for avaluo in avaluos:
            small_style = getSampleStyleSheet()
            custom_style = ParagraphStyle(
                'custom_style',
                parent=small_style['BodyText'],
                alignment=1,
                fontSize=8,  # Adjust the font size as needed
            )
            folio = avaluo['tipo']  + '-' + avaluo['cliente'] + '/' + '^^' + avaluo['dtcreate'][5:7] + '-' + avaluo['dtcreate'][2:4] + '/' + '^^' +avaluo['consecutivo'] + '-' + avaluo['valuador']
            folio = Paragraph(folio.replace('^^', '<br />'), custom_style)#getSampleStyleSheet()['BodyText'])
            #ubicacion = Paragraph(avaluo['ubicacion'].replace(' ', '<br />'), getSampleStyleSheet()['BodyText'])
            estado = max_words_per_line(col_widths3[1], 12, avaluo['estado'])
            #print(estado)
            estado = ' '.join([str(item) for item in estado])

            municipio = max_words_per_line(col_widths3[1], 12, avaluo['municipio'])
            municipio = ' '.join([str(item) for item in municipio])

            colonia = max_words_per_line(col_widths3[1], 12, avaluo['colonia'])
            colonia = ' '.join([str(item) for item in colonia])
            
            calle = max_words_per_line(col_widths3[1], 12, avaluo['calle'])
            calle = ' '.join([str(item) for item in calle])
            
            ubicacion =  estado + ',' + '^^' + municipio + ',' + '^^' + colonia + ',' + '^^' + calle +' '+ avaluo['numero'] + '^^' + 'Lote: '+ avaluo['lote']#avaluo['estado'] + ',' + '^^' + avaluo['municipio'] + ',' + '^^'  + avaluo['colonia']
            #estado = Paragraph(estado.replace('^^', '<br />'), getSampleStyleSheet()['BodyText'])
            if avaluo['dtcreate'] == 'None' or avaluo['dtcreate'] == "0000-00-00":
                dtcrt = "No disponible"
            else:
                dtcrt = formato_fechas(avaluo['dtcreate'])


            if avaluo['dtsolicitud'] == 'None' or avaluo['dtsolicitud'] == "0000-00-00":
                dtsol = "No disponible"
            else:
                dtsol = formato_fechas(avaluo['dtsolicitud'])
            

            if avaluo['dtvaluador'] == 'None' or avaluo['dtvaluador'] == "0000-00-00":
                dtval = "No disponible"
            else:
                dtval = formato_fechas(avaluo['dtvaluador'])


            if avaluo['dtcliente'] == 'None' or avaluo['dtcliente'] == "0000-00-00":
                dtclt = "No disponible"
            else:
                dtclt = formato_fechas(avaluo['dtcliente'])


            if avaluo['dtcobro'] == 'None' or avaluo['dtcobro'] == "0000-00-00":
                dtcbr = "No disponible"
            else:
                dtcbr = formato_fechas(avaluo['dtcobro'])
              
            
            fechas = 'Alta:' + dtcrt + '^^' +'Cotización:' + dtsol + '^^' + 'Valuador:' +  dtval + '^^' + 'Cliente:' + dtclt + '^^' + 'Cobro:' + dtcbr
            fechas = Paragraph(fechas.replace('^^', '<br />'), custom_style)#getSampleStyleSheet()['BodyText'])
            #municipio = Paragraph(municipio.replace('^^', '<br />'), getSampleStyleSheet()['BodyText'])
            
        
            ubicacion = Paragraph(ubicacion.replace('^^', '<br />'), custom_style)#getSampleStyleSheet()['BodyText'], custom_style)
            
            #ubicacion.wrapOn(p,"Helvetica", 6)
            cliente = Paragraph(avaluo['cliente'].replace(' ', '<br />'), custom_style)#getSampleStyleSheet()['BodyText'])
            estatus = Paragraph(avaluo['estatus'].replace(' ', '<br />'), custom_style)#getSampleStyleSheet()['BodyText'])

            tipoimb = Paragraph(avaluo['tipoimbid'].replace(' ', '<br />'), custom_style)
            if avaluo['valor'] == 'None':
                valor = 0
            else:
                valor = '{:,}'.format(float(avaluo['valor']))
            
            data3.append([avaluo['id'],ubicacion,fechas,cliente,estatus,folio,tipoimb,valor])

        

        style = [('BACKGROUND', (0, 0), (-1, 0), fill_color),
                            ('TEXTCOLOR', (0, 0), (-1, 0), 'black'),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                            ('INNERGRID', (0, 0), (-1, -1), 0.25, 'black'),
                            ('BOX', (0, 0), (-1, -1), 0.25, 'black'),
                            ('GRID',(0,0),(-1,-1),0.5,colors.black),
                            ('SPAN', (1, 0), (1, 0))]
        # Add styles for alternating grey colors and border texture
        page_size = letter[1]
        count = 0
        count_sec = [0]
        num_pages = 0
        data_global = []
        style_list = []
        h_style = getSampleStyleSheet()
        head_style = ParagraphStyle(
            'head_style',
            parent=h_style['BodyText'],
            alignment=1,
            )
        for i in range(1, len(data3)):
            style_index = i - count
            if i % 2 == 1:  # Alternating rows
                style += ([('BACKGROUND', (0, style_index), (-1, style_index), Color(0.9, 0.9, 0.9)),
                               ('BOX', (0, style_index), (-1, style_index), 0.25, 'black'),
                               ('BOTTOMPADDING', (0, style_index), (-1, style_index), 5)])
            if i % 2 == 0:  # Alternating rows
                style += ([('BACKGROUND', (0, style_index), (-1, style_index), Color(0.8, 0.8, 0.8)),
                            ('BOX', (0, style_index), (-1, style_index), 0.25, 'black'),
                            ('BOTTOMPADDING', (0, style_index), (-1, style_index), 5)])
            
            head_tipoimb = Paragraph('Tipo Inmueble'.replace(' ', '<br />'), head_style)
            new_data = [['Id','Dirección', 'Fechas','Cliente','Estatus', 'Folio',head_tipoimb,'Valor']]+data3[count:i]
            table3 = Table(new_data, colWidths=col_widths3)
            table3.setStyle(style)
            w, h = table3.wrap(0, 0)

            # Check if adding the table to the current page exceeds the available space
            if num_pages == 0:
                logic_condition = h > (page_size - 510)
            else:
                logic_condition = h > (page_size - 370)
            if logic_condition:
                # Start a new page
                #p.showPage()
                num_pages += 1
                #current_page_height = 0 
                count = i
                count_sec.append(i)
                style_list.append(style)
                data_global.append(new_data)
                style = [('BACKGROUND', (0, 0), (-1, 0), fill_color),
                            ('TEXTCOLOR', (0, 0), (-1, 0), 'black'),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                            ('INNERGRID', (0, 0), (-1, -1), 0.25, 'black'),
                            ('BOX', (0, 0), (-1, -1), 0.25, 'black'),
                            ('GRID',(0,0),(-1,-1),0.5,colors.black),
                            ('SPAN', (1, 0), (1, 0))]
                
            #else:
            #    current_page_height = h  # Update the current page height
        head_tipoimb = Paragraph('Tipo Inmueble'.replace(' ', '<br />'), head_style)
        new_data = [['Id','Dirección', 'Fechas','Cliente','Estatus', 'Folio',head_tipoimb,'Valor']]+data3[count:]
        data_global.append(new_data)
        style = [('BACKGROUND', (0, 0), (-1, 0), fill_color),
                                    ('TEXTCOLOR', (0, 0), (-1, 0), 'black'),
                                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                    ('INNERGRID', (0, 0), (-1, -1), 0.25, 'black'),
                                    ('BOX', (0, 0), (-1, -1), 0.25, 'black'),
                                    ('GRID',(0,0),(-1,-1),0.5,colors.black),
                                    ('SPAN', (1, 0), (1, 0))]
        for i in range(1, len(data_global[-1])):
       
            if i % 2 == 1:  # Alternating rows
                style += ([('BACKGROUND', (0, i), (-1, i), Color(0.9, 0.9, 0.9)),
                               ('BOX', (0, i), (-1, i), 0.25, 'black'),
                               ('BOTTOMPADDING', (0, i), (-1, i), 5)])
            if i % 2 == 0:  # Alternating rows
                style += ([('BACKGROUND', (0,i), (-1, i), Color(0.8, 0.8, 0.8)),
                            ('BOX', (0, i), (-1, i), 0.25, 'black'),
                            ('BOTTOMPADDING', (0, i), (-1, i), 5)])
        style_list.append(style)
        if num_pages < 1:
            #print(len(data3))
            #print(len(style))
            data_global.append(data3)
            style_list.append(style)
            table3 = Table(data_global[0], colWidths=col_widths3)
            table3.setStyle(style_list[0]) #[count_sec[0]:3*(count_sec[1]+1)])
            w, h = table3.wrap(0, 0)

            table3.wrapOn(p, 800,800)
            x_coord = inch-115
            table3.drawOn(p, x_coord, 610 - h)
            p.setFillColor("black")
            p.setFont("Helvetica", 8)
            footer1 = "Av. Río Mixcoac 88-201. Col. Actipan del Valle, Alcaldía Benito Juárez, Ciudad de México 03100 Teléfonos 5662-1573 "
            p.drawString(inch - 90, letter[1]/2-210, footer1)
            footer2 = "E-mail:info@aep.com.mx"
            p.drawString(inch - 90, letter[1]/2-220, footer2)
            footer3 = "5663-4892"
            p.drawString(inch + 293, letter[1]/2-220, footer3)
            tot_pags = num_pages + 1
            current_page = i+1
            current_page = 'página' + '' + '1' + '/' + str(tot_pags)
            p.drawString(inch + 400, letter[1]/2-220, current_page)
            today = date.today()
            # dd/mm/YY
            d1 = today.strftime("%d/%m/%Y")
            p.drawString(inch + 400, letter[1]/2-210, d1)

            p.showPage()

        else:
            for i in range(num_pages+1):

                table3 = Table(data_global[i], colWidths=col_widths3)
                table3.setStyle(style_list[i]) 
                w, h = table3.wrap(0, 0)
            
                table3.wrapOn(p, 800,800)
                
                if i==0:
                    
                    y_coord = 610 - h
    
                else:

                    p.drawImage(image_path, inch - 20, letter[1] -inch -20, width=100, height=70)

                    #p.setFont("Helvetica-Bold", 12)

                    # Add the text "Hello World" below the image
                    #p.drawString(inch + 80, letter[1] - inch + 35, "Avalúos, Proyectos y Servicios")

                    p.setFont("Helvetica", 12)

                    p.drawString(inch + 80, letter[1] - inch + 15 , "www.aep.com.mx")

                    line_start = inch - 20
                    line_end = line_start + 510 # Adjust based on text width
                    line_y = letter[1] - inch - 20  # Adjust based on the position of the text
                    p.line(line_start, line_y, line_end, line_y)

                    fill_color_rgb = (207/238, 197/238, 238/238)
                    fill_color = Color(*fill_color_rgb)
                    p.setFillColor(fill_color)

                    

                    # Move to a new line before adding the table
                    p.translate(inch, -2 * inch)
                    
                    #x_coord = inch-100
                    y_coord = 820 - h
                #p.setFont("Helvetica", 10)
                x_coord = inch-115
                table3.drawOn(p, x_coord, y_coord)
                p.setFillColor("black")
                p.setFont("Helvetica", 8)
                footer1 = "Av. Río Mixcoac 88-201. Col. Actipan del Valle, Alcaldía Benito Juárez, Ciudad de México 03100 Teléfonos 5662-1573 "
                p.drawString(inch - 90, letter[1]/2-210, footer1)
                footer2 = "E-mail:info@aep.com.mx"
                p.drawString(inch - 90, letter[1]/2-220, footer2)
                footer3 = "5663-4892"
                p.drawString(inch + 293, letter[1]/2-220, footer3)
                tot_pags = num_pages + 1
                current_page = i+1
                current_page = 'página' + '' + str(current_page) + '/' + str(tot_pags)
                p.drawString(inch + 400, letter[1]/2-220, current_page)
                today = date.today()
                # dd/mm/YY
                d1 = today.strftime("%d/%m/%Y")
                p.drawString(inch + 400, letter[1]/2-210, d1)

                p.showPage()

        
        p.save()

        return response


def generar_excel(request, cliente_id, tipo_id, valuador_id, estatus_id, estado_id, municipio_id, colonia_id,dtcreate_inicial, dtcreate_final,dtsolicitud_inicial, dtsolicitud_final, dtvaluador_inicial, dtvaluador_final, dtcliente_inicial, dtcliente_final, dtcobro_inicial, dtcobro_final,dtpago_inicial, dtpago_final):
    # Your dictionary data
    data = get_avaluos(request, cliente_id, tipo_id, valuador_id, estatus_id, estado_id, municipio_id, colonia_id,dtcreate_inicial, dtcreate_final,dtsolicitud_inicial, dtsolicitud_final, dtvaluador_inicial, dtvaluador_final, dtcliente_inicial, dtcliente_final, dtcobro_inicial, dtcobro_final,dtpago_inicial, dtpago_final)
    # Create a response object with appropriate Excel headers
    today = str(date.today())
    today = today[2:4]+today[5:7]+today[8:10]

    filename = today+"_reporte_control_servicios.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
    #response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #response['Content-Disposition'] = 'attachment; filename="output.xlsx"'
    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    data = json.loads(data.content)
    avaluos = data['avaluos']
    data = {
        'Id':[],
        'Dirección':[],
        'Fechas':[],
        'Cliente':[],
       #'Valuador':[],
        'Estatus':[],
        'Folio':[],
        'Tipo de inmueble':[],
        'Valor':[],
        
    }
    dtcrt = 0
    dtsol = 0
    dtval = 0
    dtclt = 0
    dtcbr = 0
    dtpgo = 0
 


    for avaluo in avaluos:
        if avaluo['dtcreate'] == 'None' or avaluo['dtcreate'] == "0000-00-00":
            dtcrt = "No disponible"
        else:
            dtcrt = formato_fechas(avaluo['dtcreate'])


        if avaluo['dtsolicitud'] == 'None' or avaluo['dtsolicitud'] == "0000-00-00":
            dtsol = "No disponible"
        else:
            dtsol = formato_fechas(avaluo['dtsolicitud'])
        

        if avaluo['dtvaluador'] == 'None' or avaluo['dtvaluador'] == "0000-00-00":
            dtval = "No disponible"
        else:
            dtval = formato_fechas(avaluo['dtvaluador'])


        if avaluo['dtcliente'] == 'None' or avaluo['dtcliente'] == "0000-00-00":
            dtclt = "No disponible"
        else:
            dtclt = formato_fechas(avaluo['dtcliente'])


        if avaluo['dtcobro'] == 'None' or avaluo['dtcobro'] == "0000-00-00":
            dtcbr = "No disponible"
        else:
            dtcbr = formato_fechas(avaluo['dtcobro'])

        if avaluo['dtpago'] == 'None' or avaluo['dtpago'] == "0000-00-00":
            dtpgo = "No disponible"
        else:
            dtpgo = formato_fechas(avaluo['dtpago'])


        data["Id"].append(avaluo["id"])
        ubicacion =  avaluo["estado"] + ',' + chr(10) + avaluo["municipio"] + ',' + chr(10)+ avaluo["colonia"] + ',' + chr(10)+ avaluo["calle"] +' '+ avaluo['numero'] + chr(10) + 'Lote: '+ avaluo['lote']
        data["Dirección"].append(ubicacion)
        fecha =  "Alta:  " + dtcrt + chr(10) + "Solicitud:  " + dtsol + chr(10) + "Valuador:  " + dtval + chr(10) + "Cobro:  " + dtcbr + chr(10)  + "Pago:  " + dtpgo

        data["Fechas"].append(fecha)
        data["Cliente"].append(avaluo["cliente"])
        #data["Valuador"].append(avaluo["valuador"])
        data["Estatus"].append(avaluo["estatus"])
        
        data["Folio"].append(avaluo['tipo']  + '-' + avaluo['cliente'] + '/' + avaluo['dtcreate'][5:7] + '-' + avaluo['dtcreate'][2:4] + '/'  +avaluo['consecutivo'] + '-' + avaluo['valuador'])
        print(avaluo)
        data["Tipo de inmueble"].append(avaluo["tipoimbid"])

        if avaluo['valor'] == 'None':
            valor = 0
        else:
            valor = '{:,}'.format(float(avaluo["valor"]))

        data["Valor"].append(valor)

    estado = Estados.objects.get(estado_id = estado_id)

    municipio = Municipios.objects.get(municipio_id = municipio_id)

    colonia = Colonias.objects.get(colonia_id = colonia_id)

    tipo = Tipos.objects.get(tipoid = tipo_id)

    cliente = Clientes.objects.get(clienteid = cliente_id)

    valuador = Valuadores.objects.get(valuadorid = valuador_id)

    estatus = Estatus.objects.get(estatusid = estatus_id)

    # Define data for the table
    params1 = {'Parámetros':['Estado','Municipio','Colonia','Tipo de servicio','Cliente','Valuador','Estatus'],
            '': [estado.nombre,municipio.nombre,colonia.nombre, tipo.display,cliente.nombre,valuador.display,estatus.nombre]}

    headers_params1 = list(params1.keys())
    for col_num, header in enumerate(headers_params1, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    # Write data to the Excel file and apply formatting
    num_rows = max(len(params1[field]) for field in headers_params1)
    for row_num in range(1, num_rows + 1):
        for col_num, field in enumerate(headers_params1, start=1):
            cell = worksheet.cell(row=row_num + 1, column=col_num, value=params1[field][row_num - 1])
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    # Write data to the Excel file and apply formatting
    num_rows = max(len(params1[field]) for field in headers_params1)
    num_cols = len(list(params1.keys()))

    dtcrt_inicial = dtcreate_inicial
    dtcrt_final = dtcreate_final
    dtsol_inicial = dtsolicitud_inicial
    dtsol_final = dtsolicitud_final
    dtval_inicial = dtvaluador_inicial
    dtval_final = dtvaluador_final
    dtclt_inicial = dtcliente_inicial
    dtclt_final = dtcliente_final
    dtcbr_inicial = dtcobro_inicial
    dtcbr_final = dtcobro_final
    dtpg_inicial = dtpago_inicial
    dtpg_final = dtpago_final



    data2_dict = {'header1':['Parámetros','',''],
    'header2':['Fecha','Inicio','Fin']}


    log_list = [((dtcreate_inicial == '0000-00-00') or (dtcreate_inicial == '')),((dtcreate_final == '0000-00-00') or (dtcreate_final == '')),((dtsolicitud_inicial == '0000-00-00') or (dtsolicitud_inicial == '')),( (dtsolicitud_final == '0000-00-00') or (dtsolicitud_final == '')),((dtvaluador_inicial == '0000-00-00' or dtvaluador_inicial == '')),((dtvaluador_final == '0000-00-00' or dtvaluador_final == '')),((dtcliente_inicial == '0000-00-00' or dtcliente_inicial == '')),((dtcliente_final == '0000-00-00' or dtcliente_final == '')),((dtcobro_inicial == '0000-00-00' or dtcobro_inicial == ''))and((dtcobro_final == '0000-00-00' or dtcobro_final == '')),((dtpago_inicial == '0000-00-00' or dtpago_inicial == '')),((dtpago_final == '0000-00-00' or dtpago_final == ''))]

    log_val = False

    for x in log_list:
        log_val = log_val or not x
    
    
    params2 = {'Parámetros':['Fecha'],
                '': ['Inicio'],
                '.': ['Fin']}
    if log_val == True:
        count_cols = 0
        if not ((dtcreate_inicial == '0000-00-00') or (dtcreate_inicial == '')and(dtcreate_final == '0000-00-00') or (dtcreate_final == '')):
            count_cols += 1
            params2['Parámetros'].append('Alta')
            params2[''].append('')
            params2['.'].append('')
            if not ((dtcreate_inicial == '0000-00-00') or (dtcreate_inicial == '')):
                params2[''][count_cols] = formato_fechas(dtcrt_inicial)
            if not ((dtcreate_final == '0000-00-00') or (dtcreate_final == '')):
                params2['.'][count_cols] = formato_fechas(dtcrt_final)

        if not ((dtsolicitud_inicial == '0000-00-00') or (dtsolicitud_inicial == '')and(dtsolicitud_final == '0000-00-00') or (dtsolicitud_final == '')):
            count_cols += 1
            params2['Parámetros'].append('Solicitud')
            params2[''].append('')
            params2['.'].append('')
            if not ((dtsolicitud_inicial == '0000-00-00') or (dtsolicitud_inicial == '')):
                params2[''][count_cols] = formato_fechas(dtsol_inicial)
            if not ((dtsolicitud_final == '0000-00-00') or (dtsolicitud_final == '')):
                params2['.'][count_cols] = formato_fechas(dtsol_final)

        if not ((dtvaluador_inicial == '0000-00-00') or (dtvaluador_inicial == '')and(dtvaluador_final == '0000-00-00') or (dtvaluador_final == '')):
            count_cols += 1
            params2['Parámetros'].append('Valuador')
            params2[''].append('')
            params2['.'].append('')
            if not ((dtvaluador_inicial == '0000-00-00' or dtvaluador_inicial == '')):
                params2[''][count_cols] = formato_fechas(dtval_inicial)
            if not ((dtvaluador_final == '0000-00-00' or dtvaluador_final == '')):
                params2['.'][count_cols] = formato_fechas(dtval_final)

        if not ((dtcliente_inicial == '0000-00-00') or (dtcliente_inicial == '')and(dtcliente_final == '0000-00-00') or (dtcliente_final == '')):
            count_cols += 1
            params2['Parámetros'].append('Cliente')
            params2[''].append('')
            params2['.'].append('')
            if not ((dtcliente_inicial == '0000-00-00' or dtcliente_inicial == '')):
                params2[''][count_cols] = formato_fechas(dtclt_inicial)
            if not ((dtcliente_final == '0000-00-00' or dtcliente_final == '')):
                params2['.'][count_cols] = formato_fechas(dtclt_final)

        if not ((dtcobro_inicial == '0000-00-00') or (dtcobro_inicial == '')and(dtcobro_final == '0000-00-00') or (dtcobro_final == '')):
            count_cols += 1
            params2['Parámetros'].append('Cobro')
            params2[''].append('')
            params2['.'].append('')
            if not ((dtcobro_inicial == '0000-00-00' or dtcobro_inicial == '')):
                params2[''][count_cols] = formato_fechas(dtcbr_inicial)
            if not ((dtcobro_final == '0000-00-00' or dtcobro_final == '')):
                params2['.'][count_cols] = formato_fechas(dtcbr_final)

        if not ((dtpago_inicial == '0000-00-00') or (dtpago_inicial == '')and(dtpago_final == '0000-00-00') or (dtpago_final == '')):
            count_cols += 1
            params2['Parámetros'].append('Pago')
            params2[''].append('')
            params2['.'].append('')
            if not ((dtpago_inicial == '0000-00-00' or dtpago_inicial == '')):
                params2[''][count_cols] = formato_fechas(dtpg_inicial)
            if not ((dtpago_final == '0000-00-00' or dtpago_final == '')):
                params2['.'][count_cols] = formato_fechas(dtpg_final)

        print(params2)
        #num_rows2 = max(len(params2[field]) for field in headers_params1)
        #num_cols2 = len(list(params2.keys()))
        headers_params2 = list(params2.keys())
        num_rows2 = max(len(params2[field]) for field in headers_params2)
        num_cols2 = len(list(params2.keys()))
        for col_num, header in enumerate(headers_params2, start=1):
            cell = worksheet.cell(row=1, column=col_num+num_cols2+1, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
            cell = worksheet.cell(row=2, column=col_num+num_cols2+1, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        # Write data to the Excel file and apply formatting
    # num_rows = max(len(params1[field]) for field in headers_params1)
        for row_num in range(1, num_rows2 + 1):
            for col_num, field in enumerate(headers_params2, start=1):
                cell = worksheet.cell(row=row_num + 1, column=col_num+num_cols2+1, value=params2[field][row_num - 1])
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)


    


    current_row = num_rows + 3
    # Write headers to the Excel file and apply formatting
    headers = list(data.keys())
    for col_num, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=current_row+1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    # Write data to the Excel file and apply formatting
    num_rows = max(len(data[field]) for field in headers)
    for row_num in range(1, num_rows + 1):
        for col_num, field in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_num + current_row+1, column=col_num, value=data[field][row_num - 1])
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    # Save the workbook to the response
    for col_num, column_letter in enumerate(worksheet.columns, start=1):
        worksheet.column_dimensions[column_letter[0].column_letter].width = 20
    
    workbook.save(response)
    

    return response


def quitar_comas(string):
        var = string.replace(',','')
        return var

@login_required
def nuevo_avaluo(request):
    if request.method == 'GET':       
        return render(request, 'core/nuevo_avaluo.html',{
            'AvaluoForm': AvaluoForm,
            'ComentariosForm': ComentariosForm,
            'HonorariosForm': HonorariosForm
        })
    else: 
        try:
            today = date.today()
            request.POST = request.POST.copy()
            request.POST["dtcreate"] = today
            year = str(today)
            year = year[0:4]+'-'+'01-01'
         
            cons = len(list(Avaluos.objects.filter(dtcreate__range=[year,'2100-10-10']).order_by('-avaluoid').select_related('cliente','tipo','valuador','estatus')))
            request.POST["consecutivo"] = cons
       
            if len(request.POST["dtsolicitud"]) != 0:
                request.POST["dtsolicitud"] = formato_fechas_inverso(request.POST["dtsolicitud"])
            if len(request.POST["dtvaluador"]) != 0:
                request.POST["dtvaluador"] = formato_fechas_inverso(request.POST["dtvaluador"])
            if len(request.POST["dtcliente"]) != 0:
                request.POST["dtcliente"] = formato_fechas_inverso(request.POST["dtcliente"])
            if len(request.POST["dtcobro"]) != 0:
                request.POST["dtcobro"] = formato_fechas_inverso(request.POST["dtcobro"])
            if len(request.POST["dtpago"]) != 0:
                request.POST["dtpago"] = formato_fechas_inverso(request.POST["dtpago"])

            if "m2c" in request.POST:
                request.POST["m2c"] = quitar_comas(request.POST["m2c"])

            if "m2t" in request.POST:
                request.POST["m2t"] = quitar_comas(request.POST["m2t"])

            if "valor" in request.POST:
                request.POST["valor"] = quitar_comas(request.POST["valor"])
            
            if "monto" in request.POST:
                request.POST["monto"] = quitar_comas(request.POST["monto"])
            
            
            form = AvaluoForm(request.POST,request.FILES)
            
            
            
            #nuevo_avaluo = form.save(commit=False)
            form.save()

            #form = AvaluoForm()
            #form.dtcreate = today
            #form.save()

            avaluo_id = Avaluos.objects.last().avaluoid
           
            var_coms = int(request.POST["control_comentarios"])
            var_hons = int(request.POST["control_honorarios"])
            #comentarios_request = {"comentario":[],"avaluo_id":[var]}
            

            for i in range(0,var_coms):
                num = str(i+1)
                #print(request.POST["comentario_"+num])
                #nuevo_comentario = "comentario_"+num
                #comentarios_request["comentario"].append(request.POST["comentario_"+num])
                coms = Comentarios(comentario=request.POST["comentario_"+num],avaluo_id=avaluo_id,fecha=today)
                coms.save()

            for i in range(0,var_hons):
                num = str(i+1)
                monto_sin_comas = quitar_comas(request.POST["monto_"+num])
                hons = Honorarios(razon=request.POST["razon_"+num],monto=monto_sin_comas,avaluo_id=avaluo_id)
                hons.save()
            
            
            avaluo = Avaluos.objects.raw("SELECT *,  m.nombre as municipio, c.nombre as colo, e.nombre as estado, t.descripcion as tip, cl.nombre as client, va.display as valua, t.descripcion as tipo_desc, es.nombre as est FROM avaluos as a INNER JOIN colonias AS c on c.colonia_id = a.colonia_id INNER JOIN municipios AS m on m.municipio_id = c.municipio_id INNER JOIN estados AS e on e.estado_id = m.estado_id INNER JOIN tipos AS t on t.tipoid = a.tipo_id  INNER JOIN clientes AS cl on cl.clienteid = a.cliente_id INNER JOIN valuadores AS va on va.valuadorid = a.valuador_id INNER JOIN estatus AS es on es.estatusid = a.estatus_id where avaluoid = %s; ", [avaluo_id])
            avaluo = Avaluos.objects.filter(avaluoid=avaluo_id)
            comentarios = Comentarios.objects.filter(avaluo_id= avaluo_id)
            
            honorarios = Honorarios.objects.filter(avaluo_id= avaluo_id)

            indicador_tipo = 0
            for av in avaluo:
                if av.valor:
                    av.valor = '{:,}'.format(av.valor)
                    
                else:
                    av.valor = 0
            for av in avaluo:
                if av.m2c:
                    av.m2c = '{:,}'.format(av.m2c)
                if av.m2t:
                    av.m2t = '{:,}'.format(av.m2t)
                if av.tipo.display == "SUP" or av.tipo.display == "VER":
                    indicador_tipo = 1
                
            av = Avaluos.objects.get(avaluoid=avaluo_id)
            if av.consecutivo:
                folio = av.tipo.display  + '-' + av.cliente.nombre + '/' + av.dtcreate[5:7] + '-' + av.dtcreate[2:4] + '/' + str(av.consecutivo) + '-' + av.valuador.display
            else:
                folio = av.tipo.display  + '-' + av.cliente.nombre + '/' + av.dtcreate[5:7] + '-' + av.dtcreate[2:4] + '/' + 'None' + '-' + av.valuador.display

            av = Avaluos.objects.get(avaluoid=avaluo_id)
            colonia_id = av.colonia_id
            municipio_id = Colonias.objects.get(colonia_id=colonia_id)
            municipio_id = municipio_id.municipio_id
            municipio = Municipios.objects.get(municipio_id=municipio_id)
            estado_id = municipio.estado_id
            municipio=municipio.nombre

            estado = Estados.objects.get(estado_id=estado_id)
            estado=estado.nombre

            if indicador_tipo == 1:
                return render(request,'core/detalles_mid.html',{
                'avaluo': avaluo,
                'estado': estado,
                'municipio': municipio,
                'comentarios': comentarios,
                'honorarios': honorarios,
                'folio':folio,
                'indicador': indicador_tipo,
                'AvaluoForm': AvaluoForm,
                'ComentariosForm': ComentariosForm,
                'HonorariosForm': HonorariosForm
            })
            else:
                return render(request,'core/detalles_mid.html',{
                'avaluo': avaluo,
                'estado': estado,
                'municipio': municipio,
                'comentarios': comentarios,
                'honorarios': honorarios,
                'folio': folio,
                'AvaluoForm': AvaluoForm,
                'ComentariosForm': ComentariosForm,
                'HonorariosForm': HonorariosForm
            })

        except Avaluos.DoesNotExist:
            return render(request, 'core/detalles.html',{
            'form': AvaluoForm,
            'form1': ComentariosForm,
            'form2': HonorariosForm,
            'error': 'Please provide valide data'
                })

@login_required
def nueva_colonia(request):

    if request.method == 'GET':       
        return render(request, 'core/nueva_colonia.html',{
            'ColoniaForm': ColoniaForm
        })
    else: 
        try:
            form = ColoniaForm(request.POST)
            nueva_colonia = form.save(commit=False)
            nueva_colonia.save()
            col_id = Colonias.objects.last().colonia_id
            col_render = Colonias.objects.get(colonia_id=col_id)
            mun_id = col_render.municipio_id
            col_render = col_render.nombre
            mun_render = Municipios.objects.get(municipio_id = mun_id)
            est_id = mun_render.estado_id
            mun_render = mun_render.nombre
            est_render = Estados.objects.get(estado_id = est_id)
            est_render = est_render.nombre
            cp = request.POST["cp"]
            return render(request, 'core/colonia_verificacion.html',{
            'colonia': col_render,
            'municipio': mun_render,
            'estado': est_render,
            "cp": cp
                })
        except ValueError:
            return render(request, 'core/nueva_colonia.html',{
            'form': ColoniaForm,
            'error': 'Please provide valide data'
                })
        


